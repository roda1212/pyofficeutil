import argparse
import glob
import os
import lib.Log as Log
import xlrd
import openpyxl
import pathlib
import docx
import csv

def InitArgParser() -> argparse.ArgumentParser:
    """
    引数の初期化
    """
    parser = argparse.ArgumentParser(description='OfficeファイルのGrep検索')
    parser.add_argument('word', type=str, help='検索ワード')
    parser.add_argument('target', type=str, help='対象ディレクトリ')
    parser.add_argument('-out', type=str, help='結果CSVファイル', default='grep_result.csv')
    parser.add_argument('-ignorecase', type=bool, help='大文字小文字を無視（既定）', default=True)
    return parser

def EnumOfficeFiles(target):
    xls = glob.glob(f'{target}/**/*.xls', recursive=True)
    xlsm = glob.glob(f'{target}/**/*.xlsm', recursive=True)
    xlsx = glob.glob(f'{target}/**/*.xlsx', recursive=True)
    docx = glob.glob(f'{target}/**/*.docx', recursive=True)
    return sorted(xls + xlsm + xlsx + docx)

class GrepResult:
    file = ''
    path = ''
    position = ''
    value = ''
    def __init__(self, path, page, position, value):
            self.file = os.path.basename(path)
            self.path = str(pathlib.Path(path).resolve())
            self.page = page
            self.position = position
            self.value = value

    def ToDict(self):
        return {    'file': self.file,
                    'path': self.path,
                    'sheet/paragraph': self.page,
                    'position': self.position,
                    'value': self.value }
    
    def __str__(self):
        return str(self.ToDict())

    @staticmethod
    def GetHeader():
        return [ 'file', 'path', 'sheet/paragraph', 'position', 'value' ]

def IsOldExcelFile(path):
    return path.endswith('.xls')

def IsExcelFile(path):
    return path.endswith('.xlsm') or path.endswith('.xlsx')

def IsWordFile(path):
    return path.endswith('.docm') or path.endswith('.docx')

def GrepOldExcelFile(word, path, ignorecase):
    results = []
    wb = xlrd.open_workbook(path)
    for sheet in wb.sheets():
        for col in range(sheet.ncols):
            for row in range(sheet.nrows):
                val = str(sheet.cell(row, col).value)
                if ignorecase:
                    val = val.lower()
                index = 0
                while index > -1:
                    index = val.find(word, index + 1)
                    if index > -1:
                        page = sheet.name
                        pos = f'{openpyxl.utils.get_column_letter(col + 1)}{row + 1}'
                        result = GrepResult(path, page, pos, val)
                        results.append(result)
    return results

def GrepExcelFile(word, path, ignorecase):
    results = []
    wb = openpyxl.load_workbook(path)
    for sheet in wb.worksheets:
        for col in range(sheet.max_column):
            for row in range(sheet.max_row):
                val = str(sheet.cell(row + 1, col + 1).value)
                if ignorecase:
                    val = val.lower()
                index = 0
                while index > -1:
                    index = val.find(word, index + 1)
                    if index > -1:
                        page = sheet.title
                        pos = f'{openpyxl.utils.get_column_letter(col + 1)}{row + 1}'
                        result = GrepResult(path, page, pos, val)
                        results.append(result)
    return results

def GrepWordFile(word, path, ignorecase):
    results = []
    doc = docx.Document(path)
    count = 0
    for para in doc.paragraphs:
        count += 1
        text = str(para.text)
        if ignorecase:
            text = text.lower()
        index = 0
        while index > -1:
            index = text.find(word, index + 1)
            if index > -1:
                page = count
                result = GrepResult(path, page, index, para.text)
                results.append(result)

    return results

def Grep(word, target, out, ignorecase):
    results = []
    files = EnumOfficeFiles(target)
    word = word.lower() if ignorecase else word
    for file in files:
        if IsOldExcelFile(file):
            results.extend(GrepOldExcelFile(word, file, ignorecase))
        elif IsExcelFile(file):
            results.extend(GrepExcelFile(word, file, ignorecase))
        elif IsWordFile(file):
            results.extend(GrepWordFile(word, file, ignorecase))

    Log.Info(f'{len(results)} 個 見つかりました')

    with open(out, 'w', newline='') as f:
        writer = csv.DictWriter(f, GrepResult.GetHeader())
        writer.writeheader()
        for r in results:
            writer.writerow(r.ToDict())

def Main():
    args = InitArgParser().parse_args()

    if not os.path.exists(args.target):
        Log.Error(f"ディレクトリが見つかりません（{args.pdf_file}）")
        return

    Grep(str(args.word), args.target, args.out, args.ignorecase)
    

if __name__ == '__main__':
    Main()